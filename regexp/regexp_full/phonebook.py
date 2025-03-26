import csv
import openpyxl
from openpyxl.utils import get_column_letter
import re
from pprint import pprint

# 1. Чтение данных из CSV
def read_csv_file(filename):
    with open(filename, encoding="utf-8") as f:
        reader = csv.reader(f)
        contacts_list = list(reader)
    return contacts_list

def strip_spaces(contacts):
    for contact in contacts:
        for i in range(len(contact)):
            if isinstance(contact[i], str):
                contact[i] = contact[i].strip()

def clean_data(contacts):
    for contact in contacts:
        for i in range(len(contact)):
            if isinstance(contact[i], str):
                contact[i] = re.sub(r'[^\w\d\s@+().-]', '', contact[i])

def format_name(contacts):
    for contact in contacts:
        name_parts = contact[:3]
        full_name = " ".join(name_parts).strip()
        formatted_name = full_name.split(" ")

        contact[0] = formatted_name[0] if len(formatted_name) > 0 else ''
        contact[1] = formatted_name[1] if len(formatted_name) > 1 else ''
        contact[2] = formatted_name[2] if len(formatted_name) > 2 else ''

def format_phone(contacts):
    phone_pattern = re.compile(
        r"(\+7|8)?[\s(]*(\d{3})[\s)]*[\s-]*(\d{3})[\s-]*(\d{2})[\s-]*(\d{2})(\s*\(?доб\.?\s*(\d+)\)?)?"
    )
    for contact in contacts:
        if len(contact) > 5:
            phone = contact[5]
            match = phone_pattern.search(phone)
            if match:
                formatted_phone = f"+7({match.group(2)}){match.group(3)}-{match.group(4)}-{match.group(5)}"
                try:
                    if match.group(7):
                        formatted_phone += f" доб.{match.group(8)}"
                except IndexError:
                    pass
                contact[5] = formatted_phone
            else:
                print(f"Не удалось отформатировать телефон: {phone}")

def merge_duplicates(contacts):
    unique_contacts = {}
    for contact in contacts:
        key = (contact[0], contact[1])
        if key not in unique_contacts:
            unique_contacts[key] = contact
        else:
            existing_contact = unique_contacts[key]
            for i in range(len(existing_contact)):
                if existing_contact[i] == "" and contact[i] != "":
                    existing_contact[i] = contact[i]
    return list(unique_contacts.values())

def quote_fields_with_commas(contacts):
    for contact in contacts:
        for i in range(len(contact)):
            if isinstance(contact[i], str) and "," in contact[i]:
                contact[i] = f'"{contact[i]}"'

def write_to_excel(contacts_list, filename="phonebook.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Записываем данные в ячейки
    for row_num, contact in enumerate(contacts_list, start=1):
        for col_num, cell_value in enumerate(contact, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Автоматическая ширина столбцов
    for col_num in range(1, len(contacts_list[0]) + 1):
        column_letter = get_column_letter(col_num)
        column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
        sheet.column_dimensions[column_letter].width = column_width + 2

    workbook.save(filename)
    print(f"Файл '{filename}' успешно создан.")

if __name__ == "__main__":
    csv_filename = "phonebook_raw.csv"
    excel_filename = "phonebook.xlsx"

    contacts_list = read_csv_file(csv_filename)

    strip_spaces(contacts_list)
    clean_data(contacts_list)
    format_name(contacts_list)
    format_phone(contacts_list)
    contacts_list = merge_duplicates(contacts_list)
    quote_fields_with_commas(contacts_list)

    write_to_excel(contacts_list, excel_filename)

    pprint(contacts_list)