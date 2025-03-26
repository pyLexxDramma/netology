import csv
import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
import datetime
from pprint import pprint

# 1. Логгер
def logger(path):
    def __logger(old_function):
        def new_function(*args, **kwargs):
            now = datetime.datetime.now()
            function_name = old_function.__name__
            arguments = f"args: {args}, kwargs: {kwargs}"
            try:
                result = old_function(*args, **kwargs)
            except Exception as e:
                result = f"Exception: {e}"
                raise
            finally:
                with open(path, 'a', encoding='utf-8') as log_file:
                    log_file.write(f"Date and time: {now}\n")
                    log_file.write(f"Function name: {function_name}\n")
                    log_file.write(f"Arguments: {arguments}\n")
                    log_file.write(f"Return value: {result}\n")
                    log_file.write("---------------------\n")
            return result

        return new_function

    return __logger

# 2. Чтение данных из CSV
@logger("phonebook.log")
def read_csv_file(filename):
    with open(filename, encoding="utf-8") as f:
        reader = csv.reader(f)
        contacts_list = list(reader)
    return contacts_list

# Функции обработки данных (с использованием логгера)
@logger("phonebook.log")
def strip_spaces(contacts):
    for contact in contacts:
        for i in range(len(contact)):
            if isinstance(contact[i], str):
                contact[i] = contact[i].strip()

@logger("phonebook.log")
def clean_data(contacts):
    for contact in contacts:
        for i in range(len(contact)):
            if isinstance(contact[i], str):
                contact[i] = re.sub(r'[^\w\d\s@+().-]', '', contact[i])

@logger("phonebook.log")
def format_name(contacts):
    for contact in contacts:
        name_parts = contact[:3]
        full_name = " ".join(name_parts).strip()
        formatted_name = full_name.split(" ")

        contact[0] = formatted_name[0] if len(formatted_name) > 0 else ''
        contact[1] = formatted_name[1] if len(formatted_name) > 1 else ''
        contact[2] = formatted_name[2] if len(formatted_name) > 2 else ''

@logger("phonebook.log")
def format_phone(contacts):
    phone_pattern = re.compile(
        r"(\+7|8)?[\s(]*(\d{3})[\s)]*[\s-]*(\d{3})[\s-]*(\d{2})[\s-]*(\d{2})(\s*\(?доб\.?\s*(\d+)\)?)?"
    )
    for contact in contacts:
        if len(contact) > 5:
            phone = contact[5]
            match = phone_pattern.search(phone)
            if match:
                formatted_phone = f"+7({match.group(2)}){match.group(3)}-{match.group(4)}-{match.group(5)}"
                try:
                    if match.group(7):
                        formatted_phone += f" доб.{match.group(8)}"
                except IndexError:
                    pass
                contact[5] = formatted_phone
            else:
                print(f"Не удалось отформатировать телефон: {phone}")

@logger("phonebook.log")
def merge_duplicates(contacts):
    unique_contacts = {}
    for contact in contacts:
        key = (contact[0], contact[1])
        if key not in unique_contacts:
            unique_contacts[key] = contact
        else:
            existing_contact = unique_contacts[key]
            for i in range(len(existing_contact)):
                if existing_contact[i] == "" and contact[i] != "":
                    existing_contact[i] = contact[i]
    return list(unique_contacts.values())

@logger("phonebook.log")
def quote_fields_with_commas(contacts):
    for contact in contacts:
        for i in range(len(contact)):
            if isinstance(contact[i], str) and "," in contact[i]:
                contact[i] = f'"{contact[i]}"'

# 4. Создание Excel-файла и запись данных
@logger("phonebook.log")
def write_to_excel(contacts_list, filename="phonebook.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Записываем данные в ячейки
    for row_num, contact in enumerate(contacts_list, start=1):
        for col_num, cell_value in enumerate(contact, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Автоматическая ширина столбцов
    for col_num in range(1, len(contacts_list[0]) + 1):
        column_letter = get_column_letter(col_num)
        column_width = max(len(str(cell.value)) for cell in sheet[column_letter])
        sheet.column_dimensions[column_letter].width = column_width + 2

    workbook.save(filename)
    print(f"Файл '{filename}' успешно создан.")

# 5. Основная часть программы
if __name__ == "__main__":
    csv_filename = "phonebook_raw.csv"
    excel_filename = "phonebook.xlsx"

    # Чтение данных из CSV
    contacts_list = read_csv_file(csv_filename)

    # Обработка данных
    strip_spaces(contacts_list)
    clean_data(contacts_list)
    format_name(contacts_list)
    format_phone(contacts_list)
    contacts_list = merge_duplicates(contacts_list)
    quote_fields_with_commas(contacts_list)

    # Запись данных в Excel
    write_to_excel(contacts_list, excel_filename)

    pprint(contacts_list)